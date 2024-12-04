import pandas as pd
import sys
from openpyxl.styles import PatternFill
import argparse

def check_blank(data, file_name):
    if data.iloc[:, 0].isnull().any():
        print(f"Error: '{file_name}' has blank row.")
        sys.exit(1)

def check_duplicate(data, file_name):
    if data.iloc[:, 0].duplicated().any():
        print(f"Error: '{file_name}' has duplicate object.")
        sys.exit(1)

def min_max_scale(x):
    if x.max() == x.min():
        return x * 0
    return (x - x.min()) / (x.max() - x.min())

def slope_distance(row1, row2):
    distance = 0
    for i in range(len(row1)):
        for j in range(i + 1, len(row1)):
            diff_A = row1.iloc[i] - row1.iloc[j]
            diff_B = row2.iloc[i] - row2.iloc[j]
            distance += 2 * abs(diff_A - diff_B)
    return distance

class UnionFind:
    def __init__(self, n):
        self.parent = list(range(n))
        self.rank = [0] * n

    def find(self, x):
        if self.parent[x] != x:
            self.parent[x] = self.find(self.parent[x])
        return self.parent[x]

    def union(self, x, y):
        rootX = self.find(x)
        rootY = self.find(y)
        if rootX != rootY:
            if self.rank[rootX] > self.rank[rootY]:
                self.parent[rootY] = rootX
            elif self.rank[rootX] < self.rank[rootY]:
                self.parent[rootX] = rootY
            else:
                self.parent[rootY] = rootX
                self.rank[rootX] += 1

def clusters(df, threshold):
    n = len(df)
    uf = UnionFind(n)
    for i in range(n):
        for j in range(i + 1, n):
            distance = slope_distance(df.iloc[i, 1:], df.iloc[j, 1:])
            if distance < threshold:
                uf.union(i, j)
    groups = {}
    for i in range(n):
        root = uf.find(i)
        if root not in groups:
            groups[root] = []
        groups[root].append(df.iloc[i, 0])
    return groups

def calculate_distance_matrix(df, group_order):
    n = len(group_order)
    dist_matrix = pd.DataFrame(index=group_order, columns=group_order)
    for i in range(n):
        for j in range(n):
            if i != j:
                distance = slope_distance(df.loc[group_order[i], 1:], df.loc[group_order[j], 1:])
                dist_matrix.iloc[i, j] = distance
            else:
                dist_matrix.iloc[i, j] = 0
    return dist_matrix

def save_colored_network(distance_matrix, groups, threshold, output_file):
    colors = [
        'FF9999', '99FF99', '9999FF', 'FFFF99', 'FF99FF', '99FFFF',
        'FFCC99', 'CC99FF', 'FF6666', '66FF66', '6666FF', 'FFFF66',
        'FF66FF', '66FFFF', 'FFCC66', 'CCFF99', 'CC66FF', 'FF9966',
        '66CCFF', 'FF6699'
    ]

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        distance_matrix.to_excel(writer, sheet_name='Distance Matrix', index=True)
        workbook = writer.book
        worksheet = workbook['Distance Matrix']

        color_index = 0

        for group_key, members in groups.items():
            if len(members) >= 2:
                fill = PatternFill(start_color=colors[color_index % len(colors)],
                                 end_color=colors[color_index % len(colors)], fill_type='solid')
                color_index += 1

                for member in members:
                    if member in distance_matrix.index:
                        row_idx = distance_matrix.index.get_loc(member) + 2
                        name_cell = worksheet.cell(row=row_idx, column=1)
                        name_cell.fill = fill

                        for col in distance_matrix.columns:
                            if col != member:
                                distance_value = distance_matrix.loc[member, col]
                                if distance_value < threshold:
                                    col_idx = distance_matrix.columns.get_loc(col) + 2
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    cell.fill = fill

    print(f"\nThe result has been saved to '{output_file}'")
    print("Each color represents a cluster. And the colored numbers are edges in the network.")

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Cluster analysis with automatic threshold')
    parser.add_argument('-i', '--input', required=True, help='Input CSV file')
    parser.add_argument('-o', '--output', required=True, help='Output Excel file')
    args = parser.parse_args()

    # Load data and get number of columns
    print(f"Loading data from {args.input}")
    df = pd.read_csv(args.input, header=None)
    col_num = len(df.columns)
    
    # Set threshold automatically (number of columns - 2)
    threshold = col_num - 2
    print(f"Using automatic threshold: {threshold}")

    # Check data prerequisites
    print("Checking data prerequisites...")
    check_blank(df, args.input)
    check_duplicate(df, args.input)

    # Apply min-max scaling
    print("Applying min-max scaling...")
    df.iloc[:, 1:] = df.iloc[:, 1:].apply(min_max_scale, axis=1)

    # Find clusters
    print("Finding clusters...")
    groups = clusters(df, threshold=threshold)

    # Get the distance matrix based on group order
    print("Constructing distance matrix...")
    group_order = [member for group in groups.values() for member in group]
    distance_matrix = calculate_distance_matrix(df.set_index(0), group_order)

    # Save the colored network
    print("Saving colored network...")
    save_colored_network(distance_matrix, groups, threshold=threshold, output_file=args.output)

if __name__ == "__main__":
    main()